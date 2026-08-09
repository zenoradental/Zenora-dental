import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DEEP="FF0A2E36"; TEAL="FF028090"; MINT="FF02C39A"; CARD="FFE8F2F2"; LINE="FFDDE9E9"
WHITE="FFFFFFFF"; INK="FF13343B"; MUTE="FF5A7B82"; YELLOW="FFFFFF99"
BLUE=Font(name="Arial", size=10, color="FF0000FF")   # hardcoded inputs
BLACK=Font(name="Arial", size=10, color=INK)         # formulas

thin = Side(style="thin", color=LINE)
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def style_header(ws, row, ncols, start=1):
    for c in range(start, start+ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 30

def title(ws, text, sub=""):
    ws["A1"] = text
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=DEEP)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color=MUTE)
    ws.row_dimensions[1].height = 24

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

# =====================================================================
# 1. README
# =====================================================================
ws = wb.active; ws.title = "README"
title(ws, "Zenora Dental — Sales Operations Workbook",
      "WhiteFox AI · built 6 Aug 2026 · all currency in INR (₹)")
widths(ws, {"A":3, "B":26, "C":86})

rows = [
    ("", ""),
    ("HOW TO USE THIS FILE", ""),
    ("Pipeline", "Your CRM. One row per prospect. Update 'Stage' as deals move. Everything on Dashboard reads from here."),
    ("Prospect List", "Where you build the raw list before qualifying. Move a row into Pipeline once it clears the checks."),
    ("Pricing Calculator", "Change the yellow cells to model a specific deal. Shows year-1 value and your true margin."),
    ("Revenue Model", "12-month projection. Change the yellow assumptions at the top and everything below recalculates."),
    ("Dashboard", "Read-only summary. Conversion rates, pipeline value, and where deals are dying."),
    ("", ""),
    ("COLOUR CODE", ""),
    ("Yellow cells", "Inputs — you edit these."),
    ("Blue text", "Hardcoded numbers you can change."),
    ("Black text", "Formulas — do not overwrite; they will recalculate on their own."),
    ("", ""),
    ("HEALTHY BENCHMARKS", ""),
    ("Cold email → reply", "8–15%"),
    ("Reply → demo booked", "~40%"),
    ("Demo → closed won", "25–35%. Below 20% means you are qualifying too loosely, not demoing badly."),
    ("Time to live", "7 days or less — this is a core promise, so track it."),
    ("Monthly churn", "Under 5%"),
    ("", ""),
    ("BEFORE YOU CHARGE ANYONE", ""),
    ("Read the gap report", "05_Production_Readiness_Gap_Report.md lists three blockers that must close before a paying clinic goes live."),
    ("Do not demo Analytics", "The revenue figures in that tab are hardcoded sample data. Demo Command Center instead."),
]
r = 4
for a, b in rows:
    ws.cell(row=r, column=2, value=a)
    ws.cell(row=r, column=3, value=b)
    if a and not b:
        ws.cell(row=r, column=2).font = Font(name="Arial", size=10, bold=True, color=TEAL)
    else:
        ws.cell(row=r, column=2).font = Font(name="Arial", size=10, bold=True, color=INK)
        ws.cell(row=r, column=3).font = Font(name="Arial", size=10, color=INK)
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# =====================================================================
# 2. PIPELINE  (the CRM)
# =====================================================================
ws = wb.create_sheet("Pipeline")
title(ws, "Pipeline", "One row per prospect. Update Stage as the deal moves. Dashboard reads from this sheet.")
hdrs = ["Clinic Name","Owner (Dr.)","City","Phone","Email","Source","Google Rating",
        "Reviews","Has Website?","Running Ads?","Tier Pitched","Setup (₹)","Monthly (₹)",
        "Yr-1 Value (₹)","Stage","Next Action","Next Action Date","Notes"]
HR = 4
for i, h in enumerate(hdrs, start=1):
    ws.cell(row=HR, column=i, value=h)
style_header(ws, HR, len(hdrs))
ws.freeze_panes = "A5"

example = ["Smile Craft Dental","Dr. Anjali Rao","Pune","+91 98XXXXXX21","hello@smilecraft.in",
           "Meta Ad Library",4.7,132,"No","Yes","Professional",44999,2499,None,
           "Demo Booked","Run 15-min demo, lead with Command Center","2026-08-12",
           "Running ads to a page with no booking form — lead with wasted ad spend"]
for i, v in enumerate(example, start=1):
    ws.cell(row=HR+1, column=i, value=v)

LAST = 204
for r in range(HR+1, LAST+1):
    ws.cell(row=r, column=14, value=f"=IF(L{r}=\"\",\"\",L{r}+M{r}*12)")
    for c in range(1, len(hdrs)+1):
        cell = ws.cell(row=r, column=c)
        cell.border = BOX
        cell.font = BLACK if c == 14 else Font(name="Arial", size=10, color=INK)
        cell.alignment = Alignment(vertical="top", wrap_text=(c == 18))
    for c in (12, 13, 14):
        ws.cell(row=r, column=c).number_format = '₹#,##0;(₹#,##0);-'

stages = '"Not Contacted,Contacted,Replied,Demo Booked,Demo Done,Proposal Sent,Closed Won,Closed Lost,Nurture"'
dv = DataValidation(type="list", formula1=stages, allow_blank=True)
ws.add_data_validation(dv); dv.add(f"O{HR+1}:O{LAST}")
dv2 = DataValidation(type="list", formula1='"Starter,Professional,Clinic Plus"', allow_blank=True)
ws.add_data_validation(dv2); dv2.add(f"K{HR+1}:K{LAST}")
dv3 = DataValidation(type="list", formula1='"Yes,No,Bad one"', allow_blank=True)
ws.add_data_validation(dv3); dv3.add(f"I{HR+1}:I{LAST}")
dv4 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws.add_data_validation(dv4); dv4.add(f"J{HR+1}:J{LAST}")

widths(ws, {"A":22,"B":18,"C":12,"D":16,"E":24,"F":18,"G":12,"H":9,"I":13,"J":13,
            "K":15,"L":13,"M":13,"N":15,"O":16,"P":34,"Q":16,"R":42})
ws.cell(row=3, column=1, value="↓ Row 5 is a filled-in example. Overwrite it with your first real prospect.")
ws.cell(row=3, column=1).font = Font(name="Arial", size=9, italic=True, color=MUTE)

# =====================================================================
# 3. PROSPECT LIST
# =====================================================================
ws = wb.create_sheet("Prospect List")
title(ws, "Prospect List — raw sourcing",
      "Build the list here. A row that scores 3+ qualifying marks is ready to move into Pipeline.")
hdrs = ["Clinic Name","City / Area","Source","Google Rating","Reviews","Website Status",
        "Booking Form?","Running Ads?","Single Owner?","Qualify Score","Ready?","Found On"]
HR = 4
for i, h in enumerate(hdrs, start=1):
    ws.cell(row=HR, column=i, value=h)
style_header(ws, HR, len(hdrs))
ws.freeze_panes = "A5"

ex = ["Smile Craft Dental","Pune – Kothrud","Meta Ad Library",4.7,132,
      "Facebook page only","No","Yes","Yes",None,None,"2026-08-06"]
for i, v in enumerate(ex, start=1):
    ws.cell(row=HR+1, column=i, value=v)

LAST2 = 304
for r in range(HR+1, LAST2+1):
    ws.cell(row=r, column=10,
            value=f'=IF(A{r}="","",(IF(D{r}>=4,1,0))+(IF(E{r}>=20,1,0))+(IF(G{r}="No",1,0))+(IF(H{r}="Yes",1,0))+(IF(I{r}="Yes",1,0)))')
    ws.cell(row=r, column=11,
            value=f'=IF(A{r}="","",IF(J{r}>=3,"MOVE TO PIPELINE","Keep looking"))')
    for c in range(1, len(hdrs)+1):
        cell = ws.cell(row=r, column=c)
        cell.border = BOX
        cell.font = BLACK if c in (10, 11) else Font(name="Arial", size=10, color=INK)
        cell.alignment = Alignment(vertical="top")
    ws.cell(row=r, column=10).alignment = Alignment(horizontal="center")

dvq = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws.add_data_validation(dvq); dvq.add(f"G{HR+1}:I{LAST2}")
widths(ws, {"A":26,"B":20,"C":20,"D":13,"E":10,"F":24,"G":14,"H":14,"I":14,"J":14,"K":22,"L":13})
ws.cell(row=3, column=1, value="Qualify Score: +1 each for rating ≥4 · reviews ≥20 · no booking form · running ads · single owner. 3+ = worth contacting.")
ws.cell(row=3, column=1).font = Font(name="Arial", size=9, italic=True, color=MUTE)

# =====================================================================
# 4. PRICING CALCULATOR
# =====================================================================
ws = wb.create_sheet("Pricing Calculator")
title(ws, "Pricing Calculator", "Edit the yellow cells to model one specific deal.")
widths(ws, {"A":3,"B":34,"C":18,"D":52})

def block(r, label):
    ws.cell(row=r, column=2, value=label)
    ws.cell(row=r, column=2).font = Font(name="Arial", size=11, bold=True, color=TEAL)

def line(r, label, val, fmt=None, inp=False, note="", formula=False):
    ws.cell(row=r, column=2, value=label).font = Font(name="Arial", size=10, color=INK)
    c = ws.cell(row=r, column=3, value=val)
    c.font = BLACK if formula else (BLUE if inp else Font(name="Arial", size=10, color=INK))
    if inp:
        c.fill = PatternFill("solid", fgColor=YELLOW)
    if fmt:
        c.number_format = fmt
    c.border = BOX
    if note:
        n = ws.cell(row=r, column=4, value=note)
        n.font = Font(name="Arial", size=9, italic=True, color=MUTE)

block(4, "DEAL INPUTS")
line(5,  "Setup fee (₹)",            44999, '₹#,##0', inp=True,  note="Professional tier list price")
line(6,  "Monthly fee (₹)",           2499, '₹#,##0', inp=True,  note="Professional tier list price")
line(7,  "Setup discount (%)",           0, '0.0%',   inp=True,  note="Founding client = 0.40. Never discount the monthly.")
line(8,  "Annual prepay?  (1=yes)",      0, '0',      inp=True,  note="1 = client pays 12 months up front, gets 2 free")
line(9,  "Contract length (months)",    12, '0',      inp=True,  note="")

block(11, "REVENUE")
line(12, "Net setup fee",     "=C5*(1-C7)",                      '₹#,##0', formula=True)
line(13, "Months billed",     "=IF(C8=1,10,C9)",                 '0',      formula=True, note="Annual prepay bills 10 months for 12 months of service")
line(14, "Subscription revenue", "=C6*C13",                      '₹#,##0', formula=True)
line(15, "TOTAL CONTRACT VALUE", "=C12+C14",                     '₹#,##0', formula=True)
ws.cell(row=15, column=2).font = Font(name="Arial", size=11, bold=True, color=DEEP)
ws.cell(row=15, column=3).font = Font(name="Arial", size=11, bold=True, color=DEEP)
ws.cell(row=15, column=3).fill = PatternFill("solid", fgColor=CARD)

block(17, "YOUR COSTS")
line(18, "Hosting per clinic / month (₹)",  200, '₹#,##0', inp=True, note="Vercel + Atlas + Resend, amortised across clinics on shared tiers")
line(19, "Domain per year (₹)",            1200, '₹#,##0', inp=True, note="Only if you include it — Professional and Clinic Plus do")
line(20, "Your build hours",                 14, '0',      inp=True, note="Realistic for a per-clinic deploy today, per the gap report")
line(21, "Your hourly rate (₹)",           1500, '₹#,##0', inp=True, note="What your time is actually worth")
line(22, "Support hours / month",           1.5, '0.0',    inp=True, note="Budget more for month 1")

line(24, "Build cost",            "=C20*C21",                    '₹#,##0', formula=True)
line(25, "Hosting over term",     "=C18*C9",                     '₹#,##0', formula=True)
line(26, "Support cost over term","=C22*C21*C9",                 '₹#,##0', formula=True)
line(27, "TOTAL COST",            "=C24+C25+C26+C19",            '₹#,##0', formula=True)
ws.cell(row=27, column=2).font = Font(name="Arial", size=11, bold=True, color=DEEP)
ws.cell(row=27, column=3).font = Font(name="Arial", size=11, bold=True, color=DEEP)

block(29, "RESULT")
line(30, "Gross profit",          "=C15-C27",                        '₹#,##0', formula=True)
line(31, "Gross margin",          "=IFERROR(C30/C15,0)",             '0.0%',   formula=True, note="Under 60% means you are discounting too hard or under-scoping the build")
line(32, "Effective hourly rate", "=IFERROR(C30/(C20+C22*C9),0)",    '₹#,##0', formula=True, note="What you actually earned per hour on this deal")
line(33, "Payback on build cost", "=IFERROR((C24-C12)/C6,0)",        '0.0',    formula=True, note="Months of subscription to recover build cost after the setup fee. Negative = paid back on day one.")
for r in (30, 31, 32):
    ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=CARD)

ws.cell(row=35, column=2, value="Assumption: hosting cost is per-clinic and assumes several clinics share one Vercel Pro / MongoDB Atlas M10 instance.")
ws.cell(row=35, column=2).font = Font(name="Arial", size=9, italic=True, color=MUTE)
ws.cell(row=36, column=2, value="One clinic alone on paid tiers costs roughly ₹3,000/month and is close to break-even at the Professional price. See Sales Playbook §5.")
ws.cell(row=36, column=2).font = Font(name="Arial", size=9, italic=True, color=MUTE)

# =====================================================================
# 5. REVENUE MODEL
# =====================================================================
ws = wb.create_sheet("Revenue Model")
title(ws, "12-Month Revenue Model", "Edit the yellow assumptions. Everything below recalculates.")
widths(ws, {"A":3,"B":30,"C":13})
for i in range(4, 17):
    ws.column_dimensions[get_column_letter(i)].width = 12

ws.cell(row=4, column=2, value="ASSUMPTIONS").font = Font(name="Arial", size=11, bold=True, color=TEAL)
assum = [
    ("New clients per month",      2,     '0'),
    ("Avg setup fee (₹)",          44999, '₹#,##0'),
    ("Avg monthly fee (₹)",        2499,  '₹#,##0'),
    ("Monthly churn rate",         0.03,  '0.0%'),
    ("Hosting cost per client (₹)",200,   '₹#,##0'),
]
r = 5
for lab, val, fmt in assum:
    ws.cell(row=r, column=2, value=lab).font = Font(name="Arial", size=10, color=INK)
    c = ws.cell(row=r, column=3, value=val)
    c.font = BLUE; c.fill = PatternFill("solid", fgColor=YELLOW)
    c.number_format = fmt; c.border = BOX
    r += 1

HR = 12
labels = ["Month","New clients","Churned clients","Active clients","Setup revenue (₹)",
          "Recurring revenue (₹)","Total revenue (₹)","Hosting cost (₹)","Gross profit (₹)","MRR (₹)","ARR run-rate (₹)"]
for i, lab in enumerate(labels):
    ws.cell(row=HR+i, column=2, value=lab)
    ws.cell(row=HR+i, column=2).font = Font(name="Arial", size=10, bold=True, color=WHITE)
    ws.cell(row=HR+i, column=2).fill = PatternFill("solid", fgColor=TEAL)
    ws.cell(row=HR+i, column=2).border = BOX

for m in range(1, 13):
    col = 2 + m
    L = get_column_letter(col)
    P = get_column_letter(col-1)
    ws.cell(row=HR,   column=col, value=f"M{m}")
    ws.cell(row=HR,   column=col).font = Font(name="Arial", size=10, bold=True, color=WHITE)
    ws.cell(row=HR,   column=col).fill = PatternFill("solid", fgColor=TEAL)
    ws.cell(row=HR,   column=col).alignment = Alignment(horizontal="center")

    ws.cell(row=HR+1, column=col, value="=$C$5")
    ws.cell(row=HR+2, column=col, value=(0 if m == 1 else f"=ROUND({P}{HR+3}*$C$8,1)"))
    ws.cell(row=HR+3, column=col, value=(f"={L}{HR+1}" if m == 1 else f"={P}{HR+3}+{L}{HR+1}-{L}{HR+2}"))
    ws.cell(row=HR+4, column=col, value=f"={L}{HR+1}*$C$6")
    ws.cell(row=HR+5, column=col, value=f"={L}{HR+3}*$C$7")
    ws.cell(row=HR+6, column=col, value=f"={L}{HR+4}+{L}{HR+5}")
    ws.cell(row=HR+7, column=col, value=f"={L}{HR+3}*$C$9")
    ws.cell(row=HR+8, column=col, value=f"={L}{HR+6}-{L}{HR+7}")
    ws.cell(row=HR+9, column=col, value=f"={L}{HR+5}")
    ws.cell(row=HR+10,column=col, value=f"={L}{HR+9}*12")

    for i in range(1, 11):
        c = ws.cell(row=HR+i, column=col)
        c.font = BLACK; c.border = BOX
        c.number_format = ('0.0' if i in (1, 2, 3) else '₹#,##0;(₹#,##0);-')

ws.cell(row=HR+12, column=2, value="Year 1 totals").font = Font(name="Arial", size=11, bold=True, color=DEEP)
tot = [("Total revenue", HR+6), ("Total gross profit", HR+8)]
rr = HR+13
for lab, srow in tot:
    ws.cell(row=rr, column=2, value=lab).font = Font(name="Arial", size=10, bold=True, color=INK)
    c = ws.cell(row=rr, column=3, value=f"=SUM(C{srow}:N{srow})")
    c.font = BLACK; c.number_format = '₹#,##0'; c.border = BOX
    c.fill = PatternFill("solid", fgColor=CARD)
    rr += 1
ws.cell(row=rr, column=2, value="Exit MRR (month 12)").font = Font(name="Arial", size=10, bold=True, color=INK)
c = ws.cell(row=rr, column=3, value=f"=N{HR+9}")
c.font = BLACK; c.number_format = '₹#,##0'; c.border = BOX; c.fill = PatternFill("solid", fgColor=CARD)
ws.cell(row=rr+1, column=2, value="Exit ARR run-rate").font = Font(name="Arial", size=10, bold=True, color=INK)
c = ws.cell(row=rr+1, column=3, value=f"=N{HR+10}")
c.font = BLACK; c.number_format = '₹#,##0'; c.border = BOX; c.fill = PatternFill("solid", fgColor=CARD)

ws.cell(row=rr+3, column=2, value="Note: at roughly 8 active clients, deploy-per-clinic ops stop scaling. Gap report §1.4 covers the multi-tenant migration.")
ws.cell(row=rr+3, column=2).font = Font(name="Arial", size=9, italic=True, color=MUTE)

# =====================================================================
# 6. DASHBOARD
# =====================================================================
ws = wb.create_sheet("Dashboard")
title(ws, "Dashboard", "Read-only. All figures pull from the Pipeline sheet.")
widths(ws, {"A":3,"B":30,"C":18,"D":52})

ws.cell(row=4, column=2, value="PIPELINE BY STAGE").font = Font(name="Arial", size=11, bold=True, color=TEAL)
stage_list = ["Not Contacted","Contacted","Replied","Demo Booked","Demo Done",
              "Proposal Sent","Closed Won","Closed Lost","Nurture"]
r = 5
for st in stage_list:
    ws.cell(row=r, column=2, value=st).font = Font(name="Arial", size=10, color=INK)
    c = ws.cell(row=r, column=3, value=f'=COUNTIF(Pipeline!$O$5:$O$204,"{st}")')
    c.font = BLACK; c.border = BOX; c.alignment = Alignment(horizontal="center")
    r += 1
ws.cell(row=r, column=2, value="Total prospects").font = Font(name="Arial", size=10, bold=True, color=INK)
c = ws.cell(row=r, column=3, value='=COUNTA(Pipeline!$A$5:$A$204)')
c.font = BLACK; c.border = BOX; c.fill = PatternFill("solid", fgColor=CARD)
c.alignment = Alignment(horizontal="center")
TOTAL_ROW = r

ws.cell(row=16, column=2, value="CONVERSION").font = Font(name="Arial", size=11, bold=True, color=TEAL)
conv = [
    ("Contacted → Replied", '=IFERROR((C7+C8+C9+C10+C11)/(C6+C7+C8+C9+C10+C11+C12),0)',
     "Healthy: 8–15% on cold email"),
    ("Replied → Demo booked", '=IFERROR((C8+C9+C10+C11)/(C7+C8+C9+C10+C11+C12),0)',
     "Healthy: ~40%"),
    ("Demo → Closed won", '=IFERROR(C11/(C9+C10+C11+C12),0)',
     "Healthy: 25–35%. Below 20% = qualify harder, per Playbook §2"),
    ("Overall win rate", f'=IFERROR(C11/C{TOTAL_ROW},0)', "Prospects contacted that became clients"),
]
r = 17
for lab, f, note in conv:
    ws.cell(row=r, column=2, value=lab).font = Font(name="Arial", size=10, color=INK)
    c = ws.cell(row=r, column=3, value=f)
    c.font = BLACK; c.number_format = '0.0%'; c.border = BOX
    c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=4, value=note).font = Font(name="Arial", size=9, italic=True, color=MUTE)
    r += 1

ws.cell(row=23, column=2, value="VALUE").font = Font(name="Arial", size=11, bold=True, color=TEAL)
val = [
    ("Closed-won setup revenue", '=SUMIF(Pipeline!$O$5:$O$204,"Closed Won",Pipeline!$L$5:$L$204)', '₹#,##0', "Cash collected on setup fees"),
    ("Closed-won MRR",           '=SUMIF(Pipeline!$O$5:$O$204,"Closed Won",Pipeline!$M$5:$M$204)', '₹#,##0', "Recurring revenue you have won"),
    ("Closed-won ARR run-rate",  '=C24*12', '₹#,##0', "MRR × 12"),
    ("Closed-won year-1 value",  '=SUMIF(Pipeline!$O$5:$O$204,"Closed Won",Pipeline!$N$5:$N$204)', '₹#,##0', "Setup + 12 months"),
    ("Open pipeline value",      '=SUMIFS(Pipeline!$N$5:$N$204,Pipeline!$O$5:$O$204,"<>Closed Won",Pipeline!$O$5:$O$204,"<>Closed Lost")', '₹#,##0', "Year-1 value of everything not yet decided"),
    ("Avg deal size (won)",      '=IFERROR(C26/C11,0)', '₹#,##0', "Target: ₹75,000"),
]
r = 24
for lab, f, fmt, note in val:
    ws.cell(row=r, column=2, value=lab).font = Font(name="Arial", size=10, color=INK)
    c = ws.cell(row=r, column=3, value=f)
    c.font = BLACK; c.number_format = fmt; c.border = BOX
    c.fill = PatternFill("solid", fgColor=CARD)
    ws.cell(row=r, column=4, value=note).font = Font(name="Arial", size=9, italic=True, color=MUTE)
    r += 1

ws.cell(row=32, column=2, value="ACTIVITY").font = Font(name="Arial", size=11, bold=True, color=TEAL)
act = [
    ("Prospects with a next action", '=COUNTA(Pipeline!$P$5:$P$204)', "Every live deal should have one"),
    ("Deals with no next action",    f'=MAX(0,C{TOTAL_ROW}-C6-C11-C12-C33)', "These are the deals that quietly die"),
]
r = 33
for lab, f, note in act:
    ws.cell(row=r, column=2, value=lab).font = Font(name="Arial", size=10, color=INK)
    c = ws.cell(row=r, column=3, value=f)
    c.font = BLACK; c.border = BOX; c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=4, value=note).font = Font(name="Arial", size=9, italic=True, color=MUTE)
    r += 1

for s in wb.worksheets:
    s.sheet_view.showGridLines = False

wb.save("/tmp/deck/Zenora_Sales_Ops.xlsx")
print("saved")
